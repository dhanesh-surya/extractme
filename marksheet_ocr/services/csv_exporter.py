"""
CSV and Excel export service for marksheet data
"""
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter


class CSVExporter:
    """Export student marksheet data to CSV and Excel formats"""
    
    def export_students_to_csv(self, students):
        """
        Export student data to CSV format
        
        Args:
            students: QuerySet or list of Student objects
            
        Returns:
            BytesIO object containing CSV data
        """
        df = self._prepare_summary_dataframe(students)
        
        # Export to CSV
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')  # utf-8-sig for Excel compatibility
        output.seek(0)
        
        return output
    
    def export_students_to_excel(self, students):
        """
        Export student data to Excel format
        
        Args:
            students: QuerySet or list of Student objects
            
        Returns:
            BytesIO object containing Excel data
        """
        df = self._prepare_summary_dataframe(students)
        
        # Export to Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Marksheet Summary')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Marksheet Summary']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                # Use get_column_letter for proper column naming (A, B, ..., Z, AA, AB, ...)
                column_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[column_letter].width = min(max_length, 50)
        
        output.seek(0)
        return output
    
    def export_detailed_csv(self, students):
        """
        Export detailed CSV with one row per student per subject
        
        Args:
            students: QuerySet or list of Student objects
            
        Returns:
            BytesIO object containing CSV data
        """
        df = self._prepare_detailed_dataframe(students)
        
        # Export to CSV
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        
        return output
    
    def export_detailed_excel(self, students):
        """
        Export detailed Excel with one row per student per subject
        
        Args:
            students: QuerySet or list of Student objects
            
        Returns:
            BytesIO object containing Excel data
        """
        df = self._prepare_detailed_dataframe(students)
        
        # Export to Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Detailed Marks')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Detailed Marks']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                # Use get_column_letter for proper column naming (A, B, ..., Z, AA, AB, ...)
                column_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[column_letter].width = min(max_length, 30)
        
        output.seek(0)
        return output
    
    def _prepare_summary_dataframe(self, students):
        """Prepare summary DataFrame with one row per student - CLEAN FORMAT"""
        rows = []
        
        for student in students:
            # Start with basic student info
            row = {
                'Roll Number': student.roll_number,
                'Student Name': student.name,
                'Father Name': student.father_name if student.father_name else '',
            }
            
            # Add enrollment number if available
            if student.enrollment_number:
                row['Enrollment Number'] = student.enrollment_number
            
            # Get all marks for this student
            marks = student.marks.all().select_related('subject').order_by('subject__code')
            
            # Group marks by subject for cleaner display
            for mark in marks:
                subject_code = mark.subject.code
                subject_name = mark.subject.name
                
                # Add subject name as context (only once per subject)
                row[f'{subject_code} - Subject'] = subject_name
                
                # Add marks in a clean format
                # Theory section
                if mark.theory_ese is not None or mark.theory_internal is not None:
                    row[f'{subject_code} - Theory ESE'] = mark.theory_ese if mark.theory_ese is not None else 0
                    row[f'{subject_code} - Theory Internal'] = mark.theory_internal if mark.theory_internal is not None else 0
                    row[f'{subject_code} - Theory Total'] = mark.get_theory_total()
                
                # Practical section
                if mark.practical_marks is not None or mark.practical_internal is not None:
                    row[f'{subject_code} - Practical'] = mark.practical_marks if mark.practical_marks is not None else 0
                    row[f'{subject_code} - Practical Int'] = mark.practical_internal if mark.practical_internal is not None else 0
                    row[f'{subject_code} - Practical Total'] = mark.get_practical_total()
                
                # Subject total
                row[f'{subject_code} - Total Marks'] = mark.get_total_marks()
            
            # Add grand totals at the end
            row['Grand Total'] = student.get_total_marks()
            row['Percentage'] = f"{student.get_percentage():.2f}%"
            row['Result'] = student.get_result_status()
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _prepare_detailed_dataframe(self, students):
        """Prepare detailed DataFrame with one row per student per subject - CLEAN FORMAT"""
        rows = []
        
        for student in students:
            marks = student.marks.all().select_related('subject').order_by('subject__code')
            
            for mark in marks:
                row = {
                    'Roll Number': student.roll_number,
                    'Student Name': student.name,
                    'Father Name': student.father_name if student.father_name else '',
                    'Subject Code': mark.subject.code,
                    'Subject Name': mark.subject.name,
                }
                
                # Only add theory marks if they exist
                if mark.theory_ese is not None or mark.theory_internal is not None:
                    row['Theory ESE'] = mark.theory_ese if mark.theory_ese is not None else 0
                    row['Theory Internal'] = mark.theory_internal if mark.theory_internal is not None else 0
                    row['Theory Total'] = mark.get_theory_total()
                
                # Only add practical marks if they exist
                if mark.practical_marks is not None or mark.practical_internal is not None:
                    row['Practical'] = mark.practical_marks if mark.practical_marks is not None else 0
                    row['Practical Internal'] = mark.practical_internal if mark.practical_internal is not None else 0
                    row['Practical Total'] = mark.get_practical_total()
                
                # Subject total and status
                row['Subject Total'] = mark.get_total_marks()
                row['Status'] = 'FAIL' if mark.is_failed() else 'PASS'
                
                rows.append(row)
        
        # Add summary row for each student
        for student in students:
            summary_row = {
                'Roll Number': student.roll_number,
                'Student Name': student.name,
                'Father Name': student.father_name if student.father_name else '',
                'Subject Code': '----',
                'Subject Name': 'GRAND TOTAL',
                'Theory ESE': '',
                'Theory Internal': '',
                'Theory Total': '',
                'Practical': '',
                'Practical Internal': '',
                'Practical Total': '',
                'Subject Total': student.get_total_marks(),
                'Status': student.get_result_status()
            }
            rows.append(summary_row)
        
        return pd.DataFrame(rows)
